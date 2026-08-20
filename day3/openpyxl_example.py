import openpyxl

wb = openpyxl.Workbook() # creating a new work book
ws = wb.active # Within the active work book, creating a new work sheet
ws.title = "Expenses" # Re-naming the sheet as Expenses

ws["A1"] = "Item"   # Adding the Header for column-1
ws["B1"] = "Cost"   # Adding the Header for column-2

expenses = [            # Taking some Dummy data as list
    ["Rent", 1200],     # Each nested list has one row of data
    ["Utilities", 150],
    ["Internet", 60],
]
for row_idx, item in enumerate(expenses, start=2):  # Access each row of data from the list with its index as row_idx
    ws.cell(row=row_idx, column=1, value=item[0]) # adding data into the work sheet by reaching specific cell
    ws.cell(row=row_idx, column=2, value=item[1])

ws["A5"] = "Total"          
ws["B5"] = "=SUM(B2:B4)"
wb.save("monthly_expenses.xlsx")

print("Excel spreadsheet created successfully!")